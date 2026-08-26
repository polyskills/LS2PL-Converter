"""Génère le fichier au format d'import avancé Pennylane (CSV, et .xlsx pour relecture)."""
from __future__ import annotations

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.converter import PENNYLANE_COLUMNS, ConversionResult


def _format_value(val) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, float):
        # Convention alignée sur les exports LightSpeed sources : séparateur
        # décimal '.', pas de séparateur de milliers.
        return f"{val:.2f}".rstrip("0").rstrip(".") if val != int(val) else str(int(val))
    return str(val)


def build_pennylane_csv(resultats: list[ConversionResult]) -> bytes:
    """Génère le fichier CSV destiné à l'import Pennylane (séparateur ';', UTF-8 avec BOM
    pour une ouverture correcte dans Excel, mêmes 17 colonnes que le gabarit d'import avancé -
    « Code analytique » n'est volontairement pas exporté, cf. core.converter.PENNYLANE_COLUMNS)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    writer.writerow(PENNYLANE_COLUMNS)
    for res in resultats:
        for ligne in res.lignes:
            writer.writerow([_format_value(ligne.get(col, "")) for col in PENNYLANE_COLUMNS])
    return ("﻿" + buf.getvalue()).encode("utf-8")


def build_pennylane_workbook(resultats: list[ConversionResult]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Pennylane"

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    for c, name in enumerate(PENNYLANE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    for res in resultats:
        for ligne in res.lignes:
            for c, name in enumerate(PENNYLANE_COLUMNS, start=1):
                val = ligne.get(name, "")
                cell = ws.cell(row=row_idx, column=c, value=val if val != "" else None)
                cell.font = Font(name="Arial", size=10)
                if name in ("Débit et/ou Crédit", "Crédit") and isinstance(val, (int, float)) and val:
                    cell.number_format = "#,##0.00"
            row_idx += 1

    # Largeurs de base par colonne (lisible pour les valeurs courtes habituelles),
    # mais élargies au contenu réel quand il dépasse - ex. « Catégorie » reprend
    # l'intitulé complet du code analytique (potentiellement long, cf.
    # core.mapping_store), une largeur fixe trop courte le tronquait visuellement
    # à l'ouverture dans Excel (même si la valeur en cellule était complète).
    largeurs_base = [10, 10, 14, 30, 26, 12, 10, 24, 14, 12, 10, 20, 14, 10, 12, 12, 12]
    plafond = 60
    for i, (name, largeur_min) in enumerate(zip(PENNYLANE_COLUMNS, largeurs_base), start=1):
        plus_long = max(
            [len(name)] + [len(str(l.get(name, ""))) for res in resultats for l in res.lignes if l.get(name) not in (None, "")],
            default=len(name),
        )
        largeur_naturelle = plus_long + 2
        lettre = get_column_letter(i)
        ws.column_dimensions[lettre].width = min(max(largeur_min, largeur_naturelle), plafond)
        if largeur_naturelle > plafond:
            for cell in ws[lettre][1:]:  # [1:] saute l'en-tête (ligne 1)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
