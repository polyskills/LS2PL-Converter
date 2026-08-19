"""
Gestion des tables de correspondance (paramétrage métier), persistées en JSON
par client dans data/clients/<client_id>/mappings.json. Toute la logique de
"comptabilité" (comptes, codes analytiques, comptes de contrepartie...) est
éditable depuis l'app, rien n'est codé en dur dans le convertisseur.

Tables :
- comptes_de_vente     : référentiel pur des comptes de vente Pennylane
                         (code + libellé), sans lien avec LightSpeed - sert
                         uniquement à proposer une liste de choix fiable
                         (menu déroulant) plutôt que de la saisie libre.
- departements         : Département LightSpeed (LightSpeed n'a pas de
                         notion de "catégorie" indépendante : la référence
                         comptable EST le département) -> Compte de vente
                         (choisi dans comptes_de_vente) + taux de TVA nominal
                         (informatif : le taux réellement appliqué à chaque
                         ligne vient du fichier LightSpeed lui-même, pas
                         d'ici).
- codes_analytiques    : référentiel pur des codes analytiques Pennylane
                         (code + description), sans lien avec compte/pdv/
                         département - sert de liste de choix à la table
                         suivante, symétrique de comptes_de_vente.
- comptes_analytiques  : (Compte comptable, Point de vente, Département)
                         -> Famille + Code analytique (choisi dans
                         codes_analytiques). Les trois critères sont
                         nécessaires : LightSpeed ne fournissant aucune
                         notion d'analytique, c'est cette combinaison qui la
                         reconstitue entièrement. Hypothèse à confirmer avec
                         un cas réel : si un cas s'avère non réductible à
                         cette combinaison figée, une résolution dynamique
                         sera nécessaire à la place.
- comptes_paiement     : Mode de paiement LightSpeed -> Compte de contrepartie (banque/caisse)
- modes_paiement_ignores : intitulés de mode de paiement (correspondance
                         EXACTE, insensible à la casse/espaces) à exclure
                         purement et simplement du bloc "Modes de paiement" -
                         aucune ligne de débit/crédit générée pour eux, à la
                         différence d'un mode non mappé qui bloque l'export.
                         Réservé aux lignes sans valeur comptable propre :
                         jamais pour écarter un montant réel dont on ne sait
                         juste pas où l'imputer (cf. compte_ecart pour ça).
- comptes_tva          : Taux de TVA -> Compte de TVA collectée
- points_de_vente      : liste des points de vente connus (code + libellé + adresse mail
                         de réception de l'export automatique, optionnelle + adresse mail
                         de résultat, optionnelle - destinataire du CSV et du récapitulatif
                         après conversion, par défaut l'adresse de réception elle-même)
- parametres           : réglages généraux (code journal, compte d'écart/report, etc.)
"""
from __future__ import annotations

import copy
import io
import json
import os

import pandas as pd

from core.client_store import client_mappings_path

EMPTY_MAPPINGS = {
    "parametres": {
        "code_journal": "VT",
        "code_pays": "FR",
        "devise": "EUR",
        "famille_categorie_analytique": "POINT_DE_VENTE",
        "compte_ecart": "471000",
        "libelle_compte_ecart": "Compte d'attente - écart de report LightSpeed",
        "tolerance_equilibrage": 0.02,
    },
    "points_de_vente": [],
    "comptes_de_vente": [],
    "departements": [],
    "codes_analytiques": [],
    "comptes_analytiques": [],
    "comptes_paiement": [],
    "modes_paiement_ignores": [],
    "comptes_tva": [],
}

# Jeu d'exemple proposé à la création d'un client (repris de la logique du
# fichier "Patch Lightspeed vers Pennylane" transmis), purement indicatif :
# à valider et corriger avec le plan comptable réel du client avant tout
# usage en production.
DEFAULT_MAPPINGS = {
    "parametres": dict(EMPTY_MAPPINGS["parametres"]),
    "points_de_vente": [
        {"code": "REST", "libelle": "RESTAURANT"},
        {"code": "BARF", "libelle": "BAR FOOD"},
        {"code": "BARS", "libelle": "BAR SOMMELLERIE"},
        {"code": "SOM", "libelle": "SOMMELLERIE"},
        {"code": "ADD", "libelle": "VENTES ADDITIONNELLES"},
        {"code": "PARIS", "libelle": "PARIS 2.0"},
    ],
    "comptes_de_vente": [
        {"compte": "70110010", "libelle_compte": "VENTES SOLIDE TVA 10%"},
        {"compte": "70110200", "libelle_compte": "VENTE LIQUIDE TVA 20%"},
    ],
    "departements": [
        {"categorie_lightspeed": "Cuisine - Entrée", "compte": "70110010", "taux_tva": "10%"},
        {"categorie_lightspeed": "Cuisine - Plat", "compte": "70110010", "taux_tva": "10%"},
        {"categorie_lightspeed": "Cuisine - Dessert", "compte": "70110010", "taux_tva": "10%"},
        {"categorie_lightspeed": "Softs", "compte": "70110010", "taux_tva": "10%"},
        {"categorie_lightspeed": "Alcool (200)", "compte": "70110200", "taux_tva": "20%"},
        {"categorie_lightspeed": "Vin et Champagne", "compte": "70110200", "taux_tva": "20%"},
    ],
    "codes_analytiques": [
        {"code_analytique": "REST", "description": "RESTAURANT"},
        {"code_analytique": "BARF", "description": "BAR FOOD"},
    ],
    "comptes_analytiques": [
        {"compte": "70110010", "point_de_vente": pdv, "categorie_lightspeed": dep, "famille": "POINT_DE_VENTE", "code_analytique": pdv}
        for pdv in ("REST", "BARF")
        for dep in ("Cuisine - Entrée", "Cuisine - Plat", "Cuisine - Dessert", "Softs")
    ] + [
        {"compte": "70110200", "point_de_vente": pdv, "categorie_lightspeed": dep, "famille": "POINT_DE_VENTE", "code_analytique": pdv}
        for pdv in ("REST", "BARF")
        for dep in ("Alcool (200)", "Vin et Champagne")
    ],
    "comptes_paiement": [
        {"mode_paiement": "Carte bleue", "compte": "511100", "libelle_compte": "Remises de cartes bancaires"},
        {"mode_paiement": "Espèces", "compte": "530000", "libelle_compte": "Caisse"},
        {"mode_paiement": "Chèque", "compte": "511200", "libelle_compte": "Chèques à encaisser"},
        {"mode_paiement": "Ticket restaurant", "compte": "511300", "libelle_compte": "Titres restaurant à encaisser"},
        {"mode_paiement": "Deliveroo", "compte": "411100", "libelle_compte": "Créances plateformes de livraison - Deliveroo"},
        {"mode_paiement": "UberEats", "compte": "411110", "libelle_compte": "Créances plateformes de livraison - UberEats"},
        {"mode_paiement": "Lightspeed Payments", "compte": "511400", "libelle_compte": "Lightspeed Payments à encaisser"},
        {"mode_paiement": "Tap to Pay sur iPhone", "compte": "511100", "libelle_compte": "Remises de cartes bancaires"},
    ],
    "comptes_tva": [
        {"taux": "5.5%", "compte": "445710", "libelle_compte": "TVA collectée 5.5%"},
        {"taux": "10%", "compte": "445711", "libelle_compte": "TVA collectée 10%"},
        {"taux": "20%", "compte": "445712", "libelle_compte": "TVA collectée 20%"},
    ],
}


def _ensure_file(client_id: str, seed: dict) -> str:
    path = client_mappings_path(client_id)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(seed, f, ensure_ascii=False, indent=2)
    return path


def load_mappings(client_id: str) -> dict:
    path = _ensure_file(client_id, EMPTY_MAPPINGS)
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Complète les clés manquantes si le fichier a été créé par une version antérieure.
    merged = copy.deepcopy(EMPTY_MAPPINGS)
    for k, v in data.items():
        merged[k] = v
    return merged


def save_mappings(client_id: str, data: dict) -> None:
    path = client_mappings_path(client_id)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def seed_with_examples(client_id: str) -> dict:
    save_mappings(client_id, copy.deepcopy(DEFAULT_MAPPINGS))
    return copy.deepcopy(DEFAULT_MAPPINGS)


def reset_to_empty(client_id: str) -> dict:
    save_mappings(client_id, copy.deepcopy(EMPTY_MAPPINGS))
    return copy.deepcopy(EMPTY_MAPPINGS)


# --- Helpers de recherche (tolérants à la casse/espaces) -------------------

def _norm_key(s: str) -> str:
    return (s or "").strip().casefold()


def find_departement(mappings: dict, categorie_lightspeed: str) -> dict | None:
    """Retrouve le mapping du département LightSpeed (= la valeur de la
    colonne « Références comptables » du fichier source) vers son compte de
    vente. Le compte, lui, doit être choisi parmi ceux du référentiel
    « Comptes de vente » (cf. find_compte_reference)."""
    target = _norm_key(categorie_lightspeed)
    for row in mappings.get("departements", []):
        if _norm_key(row.get("categorie_lightspeed", "")) == target:
            return row
    return None


def find_compte_reference(mappings: dict, compte: str) -> dict | None:
    """Retrouve un compte dans le référentiel pur « Comptes de vente »
    (utilisé pour son libellé, et comme source de vérité des comptes valides
    proposés en liste déroulante dans les autres tables)."""
    for row in mappings.get("comptes_de_vente", []):
        if row.get("compte") == compte:
            return row
    return None


def find_code_analytique(mappings: dict, compte: str, point_de_vente: str, categorie_lightspeed: str) -> dict | None:
    """LightSpeed ne fournit aucune notion d'analytique : c'est la
    combinaison (compte, point de vente, département) qui la reconstitue -
    les trois critères sont nécessaires, un même compte pouvant porter un
    code analytique différent selon le point de vente ET selon le département
    (ex. un compte de boisson peut être « Sommellerie » ou « Bar » selon le
    département d'origine, même sur un seul et même point de vente)."""
    for row in mappings.get("comptes_analytiques", []):
        if (
            row.get("compte") == compte
            and _norm_key(row.get("point_de_vente", "")) == _norm_key(point_de_vente)
            and _norm_key(row.get("categorie_lightspeed", "")) == _norm_key(categorie_lightspeed)
        ):
            return row
    return None


def find_compte_paiement(mappings: dict, mode_paiement: str) -> dict | None:
    target = _norm_key(mode_paiement)
    for row in mappings.get("comptes_paiement", []):
        if _norm_key(row.get("mode_paiement", "")) == target:
            return row
    return None


def est_mode_paiement_ignore(mappings: dict, mode_paiement: str) -> bool:
    """True si ce mode de paiement doit être totalement exclu du bloc
    Règlements (ni débit ni crédit généré) - correspondance exacte
    (insensible casse/espaces), jamais partielle : un intitulé ambigu doit
    être ajouté explicitement, pas deviné par une correspondance floue."""
    target = _norm_key(mode_paiement)
    return any(_norm_key(row.get("mode_paiement", "")) == target for row in mappings.get("modes_paiement_ignores", []))


def find_compte_tva(mappings: dict, taux: str) -> dict | None:
    for row in mappings.get("comptes_tva", []):
        if row.get("taux") == taux:
            return row
    return None


def find_client_pdv_by_email(adresse_email: str) -> tuple[str, str] | None:
    """Retrouve (client_id, code_point_de_vente) à partir de l'adresse mail
    dédiée qui a reçu un export automatique. Utilisé par le service de fetch
    mail pour identifier client et point de vente sans jamais dépendre du nom
    de fichier : chaque point de vente a sa propre adresse (voir page
    « Table de correspondance »), ce qui rend l'identification fiable même si
    LightSpeed change un jour sa convention de nommage de fichier.

    Parcourt tous les clients à chaque appel plutôt que de maintenir un index
    séparé : le nombre de clients reste faible, et ça évite tout risque
    d'index désynchronisé après une modification manuelle du référentiel."""
    from core.client_store import list_clients  # import tardif : évite un cycle

    target = _norm_key(adresse_email)
    if not target:
        return None
    for client in list_clients():
        mappings = load_mappings(client["id"])
        for pdv in mappings.get("points_de_vente", []):
            if _norm_key(pdv.get("adresse_email", "")) == target:
                return client["id"], pdv["code"]
    return None


def find_pdv(mappings: dict, code_pdv: str) -> dict | None:
    """Retrouve la fiche d'un point de vente par son code (ex. pour lire son
    adresse_resultat depuis le service de fetch mail, une fois le point de
    vente déjà identifié par find_client_pdv_by_email)."""
    for pdv in mappings.get("points_de_vente", []):
        if pdv.get("code") == code_pdv:
            return pdv
    return None


def set_pdv_adresse_email(client_id: str, code_pdv: str, adresse_email: str) -> None:
    """Renseigne l'adresse mail d'un point de vente existant, sans écraser une
    valeur déjà personnalisée (idempotent, comme ensure_points_de_vente)."""
    mappings = load_mappings(client_id)
    changed = False
    for pdv in mappings.get("points_de_vente", []):
        if pdv.get("code") == code_pdv and not pdv.get("adresse_email"):
            pdv["adresse_email"] = adresse_email
            changed = True
    if changed:
        save_mappings(client_id, mappings)


# (nom de feuille, clé dans mappings, colonnes dans l'ordre voulu) — même
# ordre que les onglets de la page Table de correspondance. Attribution
# analytique n'y figure pas : traitée à part (cf. _groupes_attribution_analytique),
# la table stockée a une ligne par département, pas une ligne par groupe.
_TABLES_EXPORT_GLOBAL = [
    ("Points de vente", "points_de_vente", ["code", "libelle", "adresse_email", "adresse_resultat", "commentaires"]),
    ("Comptes de vente PL", "comptes_de_vente", ["compte", "libelle_compte", "commentaires"]),
    ("Codes Analytique PL", "codes_analytiques", ["code_analytique", "description", "commentaires"]),
    ("Départements LS", "departements", ["categorie_lightspeed", "compte", "taux_tva", "commentaires"]),
    ("Moyens de paiements", "comptes_paiement", ["mode_paiement", "compte", "libelle_compte", "commentaires"]),
    ("Moyens paiement ignorés", "modes_paiement_ignores", ["mode_paiement", "commentaires"]),
    ("Taux de TVA", "comptes_tva", ["taux", "compte", "libelle_compte", "commentaires"]),
]

_COLONNES_ATTRIBUTION_ANALYTIQUE = ["point_de_vente", "compte", "code_analytique", "famille", "departements"]

# Largeur de colonne (caractères) au-delà de laquelle le contenu passe en
# renvoi à la ligne plutôt que de continuer à élargir la colonne indéfiniment.
_LARGEUR_COLONNE_MAX = 60


def _affichage_code_libelle(code: str, rows: list[dict], code_key: str, libelle_key: str) -> str:
    """Même logique que pages/mapping.py:_affichage_depuis_code, dupliquée
    ici pour ne pas faire dépendre core/ d'une page Streamlit : "code -
    libellé" si le libellé est retrouvé dans le référentiel fourni, sinon le
    code seul."""
    code = (code or "").strip()
    if not code:
        return code
    for r in rows:
        if (r.get(code_key) or "").strip() == code:
            libelle = (r.get(libelle_key) or "").strip()
            return f"{code} - {libelle}" if libelle else code
    return code


def _groupes_attribution_analytique(mappings: dict) -> list[dict]:
    """Regroupe comptes_analytiques (une ligne par département en stockage)
    en une ligne par groupe (point de vente, compte, code analytique,
    famille), départements associés joints dans une seule cellule — même vue
    que le tableau "Attributions existantes" affiché à l'écran (page Table
    de correspondance, onglet Attribution analytique). Exporter la table
    stockée telle quelle éclaterait chaque groupe sur autant de lignes que
    de départements, perdant la lisibilité du tableau d'origine."""
    groupes: dict[tuple[str, str, str, str], list[str]] = {}
    for a in mappings.get("comptes_analytiques", []):
        cle = (a.get("point_de_vente", ""), a.get("compte", ""), a.get("code_analytique", ""), a.get("famille", ""))
        groupes.setdefault(cle, []).append(a.get("categorie_lightspeed", ""))

    comptes_de_vente = mappings.get("comptes_de_vente", [])
    codes_analytiques = mappings.get("codes_analytiques", [])
    return [
        {
            "point_de_vente": pdv,
            "compte": _affichage_code_libelle(compte, comptes_de_vente, "compte", "libelle_compte"),
            "code_analytique": _affichage_code_libelle(code, codes_analytiques, "code_analytique", "description"),
            "famille": famille,
            "departements": ", ".join(sorted((d for d in deps if d), key=str.casefold)),
        }
        for (pdv, compte, code, famille), deps in sorted(groupes.items())
    ]


def _ajuster_largeurs_colonnes(ws, df: pd.DataFrame) -> None:
    """Largeur de chaque colonne ajustée à son contenu (en-tête compris) —
    par défaut openpyxl laisse toutes les colonnes à une largeur fixe
    identique, illisible dès qu'un champ (ex. libellé, commentaires) dépasse
    quelques caractères. Au-delà de _LARGEUR_COLONNE_MAX, la colonne
    n'est plus élargie : le contenu passe en renvoi à la ligne (wrap_text)
    à la place, pour ne jamais tronquer visuellement un contenu réellement
    plus long, seulement l'afficher sur plusieurs lignes dans la cellule."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    for i, colonne in enumerate(df.columns, start=1):
        plus_long = max(
            [len(str(colonne))] + [len(str(v)) for v in df[colonne] if v not in (None, "")],
            default=len(str(colonne)),
        )
        lettre = get_column_letter(i)
        largeur_naturelle = plus_long + 2
        ws.column_dimensions[lettre].width = min(largeur_naturelle, _LARGEUR_COLONNE_MAX)
        if largeur_naturelle > _LARGEUR_COLONNE_MAX:
            for cell in ws[lettre][1:]:  # [1:] saute l'en-tête (ligne 1)
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_export_global_xlsx(mappings: dict) -> bytes:
    """Classeur .xlsx avec un onglet par table de correspondance du
    référentiel **enregistré** (contrairement aux exports CSV par onglet de
    la page Table de correspondance, qui reflètent l'état affiché à l'écran,
    y compris non enregistré) — pratique pour un export complet en un clic
    (archivage, envoi à un tiers), sans télécharger 8 CSV séparés. Utilisé
    depuis la page Réglages > Sauvegarde. Largeur de colonnes ajustée au
    contenu de chaque onglet."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nom_feuille, cle_mappings, colonnes in _TABLES_EXPORT_GLOBAL:
            rows = mappings.get(cle_mappings, [])
            df = pd.DataFrame(rows).reindex(columns=colonnes).fillna("") if rows else pd.DataFrame(columns=colonnes)
            nom = nom_feuille[:31]  # Excel limite un nom d'onglet à 31 caractères
            df.to_excel(writer, sheet_name=nom, index=False)
            _ajuster_largeurs_colonnes(writer.sheets[nom], df)

        lignes_attribution = _groupes_attribution_analytique(mappings)
        df_attribution = (
            pd.DataFrame(lignes_attribution).reindex(columns=_COLONNES_ATTRIBUTION_ANALYTIQUE)
            if lignes_attribution else pd.DataFrame(columns=_COLONNES_ATTRIBUTION_ANALYTIQUE)
        )
        df_attribution.to_excel(writer, sheet_name="Attribution analytique", index=False)
        _ajuster_largeurs_colonnes(writer.sheets["Attribution analytique"], df_attribution)
    return buf.getvalue()
