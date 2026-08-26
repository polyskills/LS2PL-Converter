"""Tests de core.mapping_store : export global .xlsx (un onglet par table)."""
import io
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import pytest

from core.client_store import CLIENTS_DIR, create_client
from core.mapping_store import EMPTY_MAPPINGS, build_export_global_xlsx, load_mappings, save_mappings


@pytest.fixture(autouse=True)
def _clean_clients_dir():
    shutil.rmtree(CLIENTS_DIR, ignore_errors=True)
    yield
    shutil.rmtree(CLIENTS_DIR, ignore_errors=True)


def test_build_export_global_xlsx_un_onglet_par_table():
    mappings = {
        **EMPTY_MAPPINGS,
        "points_de_vente": [{"code": "REST", "libelle": "Restaurant", "adresse_email": "", "adresse_resultat": "", "commentaires": ""}],
        "comptes_de_vente": [{"compte": "70110010", "libelle_compte": "Ventes", "commentaires": ""}],
    }
    contenu = build_export_global_xlsx(mappings)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))

    assert wb.sheetnames == [
        "Points de vente",
        "Comptes de vente PL",
        "Codes Analytique PL",
        "Départements LS",
        "Moyens de paiements",
        "Moyens paiement ignorés",
        "Taux de TVA",
        "Attribution analytique",
    ]
    ws = wb["Points de vente"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("code", "libelle", "adresse_email", "adresse_resultat", "commentaires")
    assert rows[1][:2] == ("REST", "Restaurant")


def test_build_export_global_xlsx_tables_vides_gardent_les_colonnes():
    contenu = build_export_global_xlsx(EMPTY_MAPPINGS)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb["Taux de TVA"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows == [("taux", "compte", "libelle_compte", "commentaires")]  # en-tête seul, table vide


def test_build_export_global_xlsx_departements_affiche_compte_avec_libelle():
    # Le stockage ne garde que le code ("compte": "70110010") ; l'export doit
    # afficher "code - libellé" comme le menu déroulant à l'écran, pas le code seul.
    mappings = {
        **EMPTY_MAPPINGS,
        "comptes_de_vente": [{"compte": "70110010", "libelle_compte": "VENTES SOLIDE TVA 10%", "commentaires": ""}],
        "departements": [
            {"categorie_lightspeed": "Cuisine - Entrée", "compte": "70110010", "taux_tva": "10%", "commentaires": ""},
        ],
    }
    contenu = build_export_global_xlsx(mappings)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb["Départements LS"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("categorie_lightspeed", "compte", "taux_tva", "commentaires")
    assert rows[1][1] == "70110010 - VENTES SOLIDE TVA 10%"


def test_build_export_global_xlsx_departements_compte_sans_libellé_connu():
    # Compte non retrouvé dans Comptes de vente PL (référentiel incomplet ou
    # désynchronisé) : le code seul est affiché, jamais une valeur vide.
    mappings = {
        **EMPTY_MAPPINGS,
        "departements": [
            {"categorie_lightspeed": "Softs", "compte": "70110099", "taux_tva": "10%", "commentaires": ""},
        ],
    }
    contenu = build_export_global_xlsx(mappings)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb["Départements LS"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[1][1] == "70110099"


def test_build_export_global_xlsx_largeur_colonnes_ajustee_au_contenu():
    mappings = {
        **EMPTY_MAPPINGS,
        "comptes_de_vente": [
            {"compte": "70110010", "libelle_compte": "VENTES SOLIDE TVA 10% RESTAURANT PRINCIPAL", "commentaires": ""},
        ],
    }
    contenu = build_export_global_xlsx(mappings)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb["Comptes de vente PL"]

    largeur_compte = ws.column_dimensions["A"].width  # "compte" (8) vs "70110010" (8)
    largeur_libelle = ws.column_dimensions["B"].width  # "libelle_compte" (14) vs le long libellé (43)

    assert largeur_libelle > largeur_compte  # bien ajustée au contenu le plus long, pas une largeur fixe uniforme
    assert largeur_libelle < 62  # plafonnée, pas illimitée


def test_build_export_global_xlsx_largeur_plafonnee_pour_texte_tres_long():
    mappings = {
        **EMPTY_MAPPINGS,
        "comptes_de_vente": [{"compte": "1", "libelle_compte": "x", "commentaires": "y" * 500}],
    }
    contenu = build_export_global_xlsx(mappings)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb["Comptes de vente PL"]
    assert ws.column_dimensions["C"].width == 60  # plafond, pas 502
    # Le contenu réel n'est jamais tronqué : renvoi à la ligne dans la cellule plutôt que coupé.
    cellule = ws["C2"]
    assert cellule.value == "y" * 500
    assert cellule.alignment.wrap_text is True


def test_build_export_global_xlsx_attribution_analytique_groupee_pas_eclatee():
    # Reproduit le bug rapporté : un même groupe (pdv, compte, code analytique, famille)
    # avec plusieurs départements doit rester sur UNE ligne, départements réunis dans une
    # seule cellule — pas une ligne par département comme dans le stockage brut.
    mappings = {
        **EMPTY_MAPPINGS,
        "comptes_de_vente": [{"compte": "70110010", "libelle_compte": "Ventes solides", "commentaires": ""}],
        "codes_analytiques": [{"code_analytique": "REST", "description": "Restaurant", "commentaires": ""}],
        "comptes_analytiques": [
            {"point_de_vente": "REST", "compte": "70110010", "categorie_lightspeed": "Cuisine - Entrée",
             "famille": "POINT_DE_VENTE", "code_analytique": "REST", "commentaires": ""},
            {"point_de_vente": "REST", "compte": "70110010", "categorie_lightspeed": "Cuisine - Plat",
             "famille": "POINT_DE_VENTE", "code_analytique": "REST", "commentaires": ""},
            {"point_de_vente": "REST", "compte": "70110010", "categorie_lightspeed": "Cuisine - Dessert",
             "famille": "POINT_DE_VENTE", "code_analytique": "REST", "commentaires": ""},
        ],
    }
    contenu = build_export_global_xlsx(mappings)
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb["Attribution analytique"]
    rows = list(ws.iter_rows(values_only=True))

    assert rows[0] == ("point_de_vente", "compte", "code_analytique", "famille", "departements")
    assert len(rows) == 2  # en-tête + UNE seule ligne pour le groupe (pas 3)
    assert rows[1][0] == "REST"
    assert rows[1][1] == "70110010 - Ventes solides"
    # Code analytique affiché tel quel, sans concaténer la description (contrairement à "compte") :
    # le code peut déjà être la chaîne longue à afficher (ex. "ASPP - Alcools & Cocktails alcoolisés").
    assert rows[1][2] == "REST"
    assert rows[1][4] == "Cuisine - Dessert, Cuisine - Entrée, Cuisine - Plat"


def test_load_mappings_repare_un_code_analytique_historiquement_tronque():
    # Reproduit le bug rapporté : avant correction, le menu déroulant enregistrait le
    # code_analytique tronqué au premier " - " (ex. "ASPP" au lieu de "ASPP - Alcools &
    # Cocktails alcoolisés"). load_mappings doit réparer automatiquement cette donnée
    # historique dès lors qu'un seul code du référentiel partage ce préfixe.
    client = create_client("Test Réparation")
    data = {
        **EMPTY_MAPPINGS,
        "codes_analytiques": [{"code_analytique": "ASPP - Alcools & Cocktails alcoolisés", "description": "", "commentaires": ""}],
        "comptes_analytiques": [
            {"point_de_vente": "REST", "compte": "70110010", "categorie_lightspeed": "Cuisine - Entrée",
             "famille": "POINT_DE_VENTE", "code_analytique": "ASPP", "commentaires": ""},
        ],
    }
    save_mappings(client["id"], data)

    reloaded = load_mappings(client["id"])
    assert reloaded["comptes_analytiques"][0]["code_analytique"] == "ASPP - Alcools & Cocktails alcoolisés"

    # La réparation est persistée (pas seulement en mémoire) : un second chargement
    # retrouve directement le code complet sans dépendre du référentiel.
    relu_du_disque = load_mappings(client["id"])
    assert relu_du_disque["comptes_analytiques"][0]["code_analytique"] == "ASPP - Alcools & Cocktails alcoolisés"


def test_load_mappings_ne_repare_pas_si_prefixe_ambigu():
    # Deux codes différents partagent le même préfixe "ASPP" : impossible de deviner
    # lequel était visé, la réparation automatique doit s'abstenir plutôt que se tromper.
    client = create_client("Test Ambiguïté")
    data = {
        **EMPTY_MAPPINGS,
        "codes_analytiques": [
            {"code_analytique": "ASPP - Alcools & Cocktails alcoolisés", "description": "", "commentaires": ""},
            {"code_analytique": "ASPP - Softs", "description": "", "commentaires": ""},
        ],
        "comptes_analytiques": [
            {"point_de_vente": "REST", "compte": "70110010", "categorie_lightspeed": "Cuisine - Entrée",
             "famille": "POINT_DE_VENTE", "code_analytique": "ASPP", "commentaires": ""},
        ],
    }
    save_mappings(client["id"], data)

    reloaded = load_mappings(client["id"])
    assert reloaded["comptes_analytiques"][0]["code_analytique"] == "ASPP"  # inchangé, ambigu


def test_load_mappings_ne_touche_pas_un_code_deja_correct():
    client = create_client("Test Code Correct")
    data = {
        **EMPTY_MAPPINGS,
        "codes_analytiques": [{"code_analytique": "REST", "description": "Restaurant", "commentaires": ""}],
        "comptes_analytiques": [
            {"point_de_vente": "REST", "compte": "70110010", "categorie_lightspeed": "Cuisine - Entrée",
             "famille": "POINT_DE_VENTE", "code_analytique": "REST", "commentaires": ""},
        ],
    }
    save_mappings(client["id"], data)

    reloaded = load_mappings(client["id"])
    assert reloaded["comptes_analytiques"][0]["code_analytique"] == "REST"
