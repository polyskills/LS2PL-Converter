"""Tests de core.pennylane_export : le classeur .xlsx généré ne doit jamais
tronquer visuellement une colonne, notamment « Catégorie » qui reprend
l'intitulé complet du code analytique (potentiellement long)."""
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from core.converter import PENNYLANE_COLUMNS, ConversionResult
from core.pennylane_export import build_pennylane_csv, build_pennylane_workbook


def _ligne(**overrides) -> dict:
    ligne = {col: "" for col in PENNYLANE_COLUMNS}
    ligne.update(
        {
            "Date": "2026-08-01",
            "Code Journal": "VT",
            "Numéro de compte": "70110010",
            "Débit et/ou Crédit": 0,
            "Crédit": 100.0,
            "Identifiant de ligne": 1,
            "Poids analytique": 1,
        }
    )
    ligne.update(overrides)
    return ligne


def test_categorie_longue_elargit_la_colonne_dans_le_xlsx():
    long_code = "ASPP - Alcools & Cocktails alcoolisés servis en salle et en terrasse"
    res = ConversionResult(source_filename="test.csv", point_de_vente="REST")
    res.lignes.append(_ligne(Catégorie=long_code))

    contenu = build_pennylane_workbook([res])
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb["Import Pennylane"]

    idx_categorie = PENNYLANE_COLUMNS.index("Catégorie") + 1
    from openpyxl.utils import get_column_letter
    lettre = get_column_letter(idx_categorie)

    # La colonne s'élargit largement au-delà de la largeur fixe historique (14) -
    # jamais de troncature visuelle du code analytique complet.
    assert ws.column_dimensions[lettre].width > 14
    assert ws.cell(row=2, column=idx_categorie).value == long_code


def test_categorie_tres_longue_plafonnee_avec_renvoi_a_la_ligne():
    tres_long = "X" * 200
    res = ConversionResult(source_filename="test.csv", point_de_vente="REST")
    res.lignes.append(_ligne(Catégorie=tres_long))

    contenu = build_pennylane_workbook([res])
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb["Import Pennylane"]

    idx_categorie = PENNYLANE_COLUMNS.index("Catégorie") + 1
    from openpyxl.utils import get_column_letter
    lettre = get_column_letter(idx_categorie)

    assert ws.column_dimensions[lettre].width == 60  # plafond, pas 202
    cellule = ws.cell(row=2, column=idx_categorie)
    assert cellule.value == tres_long  # jamais tronqué en cellule, seulement à l'affichage
    assert cellule.alignment.wrap_text is True


def test_csv_contient_toujours_le_code_analytique_complet():
    long_code = "ASPP - Alcools & Cocktails alcoolisés"
    res = ConversionResult(source_filename="test.csv", point_de_vente="REST")
    res.lignes.append(_ligne(Catégorie=long_code))

    contenu = build_pennylane_csv([res]).decode("utf-8-sig")
    assert long_code in contenu
